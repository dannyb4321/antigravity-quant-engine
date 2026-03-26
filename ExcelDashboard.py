import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from SignalValidator import get_signal_score, get_market_data
import os

def generate_excel_dashboard():
    try:
        score, detalles, direccion = get_signal_score()
        ccl, p_local, tasa = get_market_data()
        
        # Simulación de Delta para tu C6900
        strike = 6900
        delta = 0.5 + ((p_local - strike) / strike * 1.5)
        delta = max(0.1, min(0.9, delta))

        wb = Workbook()
        ws = wb.active
        ws.title = "Bitacora Quant"
        
        # Estética
        color = "00FF00" if score >= 4 else "FFFF00"
        ws['B2'], ws['B3'] = "ESTADO", f"{direccion} ({score}/5)"
        ws['B3'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        ws['D2'], ws['D3'], ws['D4'] = "MONITOR", f"CCL: ${ccl}", f"LOCAL: ${p_local}"
        ws['F2'], ws['F3'] = "GRIEGAS C6900", f"Delta Est: {round(delta, 2)}"

        wb.save("Dashboard_Trading.xlsx")
        print(f"✅ Dashboard actualizado a las {pd.Timestamp.now().strftime('%H:%M')}")
    except PermissionError:
        print("❌ Cerrá el Excel para actualizar.")

if __name__ == "__main__":
    generate_excel_dashboard()